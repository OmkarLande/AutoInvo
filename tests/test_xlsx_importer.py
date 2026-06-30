"""Tests for the XlsxImporter."""

import io
import datetime
import pytest
from pathlib import Path

import openpyxl
from autoinvo.importers.xlsx_importer import XlsxImporter
from autoinvo.importers.models import XlsxImportOptions
from autoinvo.importers.exceptions import ImportError

@pytest.fixture
def importer() -> XlsxImporter:
    return XlsxImporter()

def test_valid_workbook(importer: XlsxImporter, tmp_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Name", "Age"])
    ws.append(["Alice", 30])
    ws.append(["Bob", 25])
    path = tmp_path / "valid.xlsx"
    wb.save(path)
    
    result = importer.parse(path)
    assert result.headers == ["Name", "Age"]
    assert result.row_count == 2
    assert result.rows[0]["Name"] == "Alice"
    assert result.rows[0]["Age"] == 30
    
def test_multiple_sheets(importer: XlsxImporter, tmp_path: Path):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.append(["A"])
    ws1.append([1])
    
    ws2 = wb.create_sheet("Sheet2")
    ws2.append(["B"])
    ws2.append([2])
    
    path = tmp_path / "multi.xlsx"
    wb.save(path)
    
    # Test index
    res1 = importer.parse(path, XlsxImportOptions(sheet_index=1))
    assert res1.headers == ["B"]
    assert res1.rows[0]["B"] == 2
    
    # Test name
    res2 = importer.parse(path, XlsxImportOptions(sheet_name="Sheet2"))
    assert res2.headers == ["B"]

def test_empty_file_invalid_workbook(importer: XlsxImporter, tmp_path: Path):
    path = tmp_path / "empty.xlsx"
    path.write_bytes(b"")
    
    with pytest.raises(ImportError, match="Invalid or empty"):
        importer.parse(path)
        
def test_hidden_sheet(importer: XlsxImporter, tmp_path: Path):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Visible"
    ws1.append(["A"])
    ws1.append([1])
    
    ws2 = wb.create_sheet("Hidden")
    ws2.sheet_state = 'hidden'
    ws2.append(["B"])
    ws2.append([2])
    
    path = tmp_path / "hidden.xlsx"
    wb.save(path)
    
    res = importer.parse(path, XlsxImportOptions(sheet_name="Hidden"))
    assert res.headers == ["B"]

def test_date_cells(importer: XlsxImporter, tmp_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Date"])
    dt = datetime.datetime(2025, 1, 1, 12, 0)
    ws.append([dt])
    path = tmp_path / "date.xlsx"
    wb.save(path)
    
    result = importer.parse(path)
    assert isinstance(result.rows[0]["Date"], datetime.datetime)
    assert result.rows[0]["Date"] == dt

def test_formula_cells(importer: XlsxImporter, tmp_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Val1", "Val2", "Sum"])
    ws.append([10, 20, "=A2+B2"])
    path = tmp_path / "formula.xlsx"
    wb.save(path)
    
    result = importer.parse(path)
    # Openpyxl doesn't evaluate formulas dynamically; if there's no cached value, data_only=True returns None.
    # The key is that it shouldn't return the formula string.
    assert result.rows[0]["Sum"] != "=A2+B2"

def test_merged_cells(importer: XlsxImporter, tmp_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["A", "B"])
    ws.append(["Merged", "Hidden"])
    ws.merge_cells('A2:B2')
    path = tmp_path / "merged.xlsx"
    wb.save(path)
    
    result = importer.parse(path)
    assert result.rows[0]["A"] == "Merged"
    # Merged cells return None in other cells, our implementation normalizes None to ""
    assert result.rows[0]["B"] == ""

def test_duplicate_headers(importer: XlsxImporter, tmp_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["A", "A"])
    ws.append([1, 2])
    path = tmp_path / "dup.xlsx"
    wb.save(path)
    
    with pytest.raises(ImportError, match="duplicate"):
        importer.parse(path)

def test_empty_headers(importer: XlsxImporter, tmp_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([None, "B"])
    ws.append([1, 2])
    path = tmp_path / "empty.xlsx"
    wb.save(path)
    
    with pytest.raises(ImportError, match="empty headers"):
        importer.parse(path)
        
def test_missing_sheet(importer: XlsxImporter, tmp_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["A"])
    ws.append([1])
    path = tmp_path / "missing.xlsx"
    wb.save(path)
    
    with pytest.raises(ImportError, match="Sheet not found"):
        importer.parse(path, XlsxImportOptions(sheet_name="Nonexistent"))
        
def test_invalid_path(importer: XlsxImporter):
    with pytest.raises(ImportError):
        importer.parse("")

def test_bytes_input(importer: XlsxImporter, tmp_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["A"])
    ws.append([1])
    path = tmp_path / "bytes.xlsx"
    wb.save(path)
    
    data = path.read_bytes()
    result = importer.parse(data)
    assert result.row_count == 1
    
def test_binary_io_input(importer: XlsxImporter, tmp_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["A"])
    ws.append([1])
    path = tmp_path / "bin.xlsx"
    wb.save(path)
    
    data = io.BytesIO(path.read_bytes())
    result = importer.parse(data)
    assert result.row_count == 1
