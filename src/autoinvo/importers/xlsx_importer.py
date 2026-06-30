"""XLSX Importer for the Data Import Engine."""

import io
from pathlib import Path

import openpyxl
from autoinvo.importers.contracts import DataImporter, ImportSource
from autoinvo.importers.models import ImportResult, XlsxImportOptions
from autoinvo.importers.exceptions import ImportError


class XlsxImporter(DataImporter):
    """Importer for Excel (XLSX) data."""

    def parse(self, source: ImportSource, options: XlsxImportOptions | None = None) -> ImportResult:
        """Parses an XLSX source into an ImportResult.
        
        Args:
            source: The XLSX input data source.
            options: Configuration options for XLSX import.
            
        Returns:
            An ImportResult containing headers and rows.
            
        Raises:
            ImportError: If parsing fails or input is invalid.
        """
        if options is None:
            options = XlsxImportOptions()

        if source is None:
            raise ImportError("Source cannot be None")

        file_obj = None
        should_close = False
        wb = None

        try:
            if isinstance(source, (str, Path)):
                path = Path(source)
                if str(path).strip() == "":
                    raise ImportError("Source path cannot be empty")
                if not path.is_file():
                    raise ImportError(f"File not found: {path}")
                file_obj = path.open('rb')
                should_close = True
            elif isinstance(source, bytes):
                if not source:
                    raise ImportError("Source bytes cannot be empty")
                file_obj = io.BytesIO(source)
                should_close = True
            elif hasattr(source, "read"):
                # BinaryIO
                file_obj = source
                should_close = False
            else:
                raise ImportError(f"Unsupported source type: {type(source)}")
            
            try:
                # Use data_only=True to evaluate formulas to values
                wb = openpyxl.load_workbook(file_obj, data_only=True)
            except Exception as e:
                raise ImportError(f"Invalid or empty workbook: {e}", original_exception=e)
                
            if not wb.sheetnames:
                raise ImportError("Workbook is empty")
                
            if options.sheet_name is not None:
                if options.sheet_name not in wb.sheetnames:
                    raise ImportError(f"Sheet not found: {options.sheet_name}")
                ws = wb[options.sheet_name]
            else:
                if options.sheet_index < 0 or options.sheet_index >= len(wb.sheetnames):
                    raise ImportError(f"Sheet index out of range: {options.sheet_index}")
                ws = wb.worksheets[options.sheet_index]
                
            # Read rows
            rows_iter = ws.iter_rows(values_only=True)
            
            try:
                raw_headers = next(rows_iter)
            except StopIteration:
                raise ImportError("Worksheet is empty")
                
            if not raw_headers or all(h is None for h in raw_headers):
                raise ImportError("Worksheet contains empty headers")
                
            # Normalize headers
            headers = []
            for h in raw_headers:
                if h is None:
                    raise ImportError("Worksheet contains empty headers")
                h_str = str(h).strip()
                if not h_str:
                    raise ImportError("Worksheet contains empty headers")
                headers.append(h_str)
                
            if len(headers) != len(set(headers)):
                raise ImportError("Worksheet contains duplicate headers")
                
            rows = []
            for row in rows_iter:
                # Skip blank rows
                if not row or all(cell is None or (isinstance(cell, str) and not str(cell).strip()) for cell in row):
                    continue
                    
                row_dict = {}
                for i, h in enumerate(headers):
                    val = row[i] if i < len(row) else None
                    # Normalize string cells by stripping them. Keep dates/numbers/bools as is.
                    if isinstance(val, str):
                        val = val.strip()
                    elif val is None:
                        val = ""
                    row_dict[h] = val
                rows.append(row_dict)
                
            return ImportResult(
                headers=headers,
                rows=rows,
                row_count=len(rows),
                source_type="xlsx"
            )
            
        except ImportError:
            raise
        except Exception as e:
            raise ImportError(f"Error parsing XLSX: {e}", original_exception=e)
        finally:
            if wb:
                wb.close()
            if should_close and file_obj:
                file_obj.close()
